import openpyxl
from pprint import pprint
import login
import re
from netmiko import ConnectHandler

# aplist = []
#
# ip_dict = {}

my_device = {'host': '10.225.26.21', 'username': login.username, 'password': login.password, 'banner_timeout' : 10 , 'conn_timeout': 20, 'fast_cli': True, 'global_delay_factor': 2, 'device_type' : 'cisco_wlc_ssh'}
net_connect = ConnectHandler(**my_device)

net_connect.send_command('config paging disable')
prompt = net_connect.find_prompt()
output1 = net_connect.send_command('show flexconnect group summary',use_textfsm = True)

y = output1.splitlines()

print (y)



# ipam = 'INMPR'
# #
# # output1 = 'INMPR-FC-ON                       24'
#
# pattern = '\s'
#
# # empty string
# replace = ' '
#
# new_string = re.sub(pattern, replace, output1)
# # pprint(new_string)
#
# y =new_string.split(' ')
#
# for i in range (len (y)):
#     # print (y[i])
#     if ipam in y[i]:
#         print (y[i])

# x = re.match(ipam, output1)
#
# if x:
#   print("Search successful.")
# else:
#   print("Search unsuccessful.")
# print (x)

# x = re.search (r'/D/D')
# y = x.findall(output1)
#
# print (y)



#
# for ap in output1:
#
#     ip_dict = {'AP_model' : ap['ap_model'], 'AP_Name' : ap['ap_name'], 'Country' : ap['country'], 'IP Address' : ap['ip'] , 'Location' : ap['location']}
#
#
#
#
#
# #print (output1)
# wb = openpyxl.load_workbook('wifi_sng.xlsx')
# sheet1 = wb.active
# #sheet2 = wb.active
# #sheet3 = wb.active
#
#
# #wb.create_sheet(index = 2, title= 'show_ip_bgp_summary')
#
# sheet1 = wb['Sheet1']
# #sheet3 = wb['show_ip_bgp_summary']
#
# x = 1
# item1 = output1[0]
#
# for k in item1.keys():
#     sheet1.cell(row=1, column=x).value = k
#     x = x + 1
#
# y = 2
# for items in output1:
#     z = 1
#     for b in items.values():
#         # if type(b) is list:
#         #     q = '//'.join(b)
#         sheet1.cell(row=y, column=z).value = b
#         # else:
#         #     sheet1.cell(row=y, column=z).value = b
#         z = z + 1
#     y = y + 1
#
#
# wb.save('wifi_sng.xlsx')
