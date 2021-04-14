import requests
from netmiko import ConnectHandler
import openpyxl
from pprint import pprint
import login
import concurrent.futures
import xlrd
import textfsm
import ntc_templates


def get_inv_solarwinds_live(uname=login.username, pasw=login.password,
                            solarwinds_url='https://solarwinds.mdlz.com:17778/SolarWinds/InformationService/v3/Json/Query?query=SELECT%20OrionNodes.NodeId,%20OrionNodes.CustomProperties.Region,%20OrionNodes.CustomProperties.Sub_Region,%20OrionNodes.CustomProperties.Country,%20OrionNodes.CustomProperties.Site,%20OrionNodes.CustomProperties.IPAM_Code,%20OrionNodes.Caption,%20OrionNodes.IP_Address,%20OrionNodes.Status,%20OrionNodes.CustomProperties.Network_Function,%20OrionNodes.CustomProperties.Site_Type,%20OrionNodes.CustomProperties.Site_Operating_Hours,%20OrionNodes.CustomProperties.Site_Metal_Rating,%20OrionNodes.CustomProperties.ManagedBy,%20OrionNodes.SysName,%20OrionNodes.Location,%20OrionNodes.Contact,%20OrionNodes.MachineType,%20OrionNodes.Vendor,%20OrionNodes.LastBoot,%20OrionNodes.IOSVersion%20From%20Orion.Nodes%20AS%20OrionNodes'):
    print('Attempting to download data from SolarWinds...')
    r = requests.get(solarwinds_url, auth=(uname, pasw), verify=False)
    if r.status_code == 200:
        print('Got data from SolarWinds!')
        jas = r.json()
        r.close()
        return jas.get('results', [])
    else :
        print (f"Something went wrong... Couldn't load data from SolarWinds. Status Code: {r.status_code}")
    return [ ]


def user_input (ipam_code):

    for item in amdb:
        if item ['IPAM_Code']== ipam_code and  'Switch' in item ['Network_Function'] and item ['Vendor'] == 'H3C' :
            ip_hcl_HP.append (item [ 'IP_Address' ])

    for item in amdb :
        if item [ 'IPAM_Code' ] == ipam_code and "Switch" in item [ 'Network_Function' ] and item ['Vendor' ] == "Cisco" :
            ip_hcl_cisco.append (item [ 'IP_Address' ])

    for item in amdb :
        if item [ 'IPAM_Code' ] == ipam_code and "Switch" in item [ 'Network_Function' ] and item ['Vendor' ] == "HPE" :
            ip_hcl_HPE.append ( item ['IP_Address'] )


    if len (ip_hcl_cisco) > 0:
        print (len(ip_hcl_cisco))
        # pprint (ip_hcl_cisco)

        with concurrent.futures.ThreadPoolExecutor(max_workers=20) as mainexecutor:
            for vip in ip_hcl_cisco :
                # VERSION(vip)
                mainexecutor.submit(VERSION,vip,ipam_code)


    # pprint(qdata)



    # print ("Below is the list of H3C Switches")
    #
    # pprint (ip_hcl_HP)
    #
    # print ('********************')
    #
    # print ("below is the list of HPE Switches")
    #
    # pprint (ip_hcl_HPE)

def VERSION (ipsw, ipam_c):
    try :
        my_device={ 'host' : ipsw,'username' : login.username,'password' : login.password,'secret' : login.password,
                    'device_type' : 'cisco_ios','global_delay_factor' : 8 }
        session=ConnectHandler (**my_device)
        prompt=session.find_prompt ()
        output=session.send_command ('show cdp neighbors detail',use_textfsm=True,expect_string=prompt)
        # pprint (output)
        for i in range ( len ( output ) ) :

            if 'AIR-AP' in output [ i ] [ 'platform' ] or  'AIR-CAP' in output [ i ] [ 'platform' ] :
                ip_dict = { 'IPAM_CODE' : ipam_c , 'WAP_IP' : output [ i ] [ 'management_ip' ] ,
                            'Local_Port' : output [ i ] [ 'local_port' ] , 'Switch_IP' : ipsw ,
                            'Device_HOSTNAME' : output [ i ] [ 'destination_host' ] ,
                            'Model_number' : output [ i ] [ 'platform' ] ,
                            'Software_Version' : output [ i ] [ 'software_version' ] }

                # pprint (ip_dict)
                ou.append ( ip_dict )

        session.disconnect ()

    except Exception as ex :
        print (f'\n\t ! Login into {ipsw} failed.',end='')
        print ('\n\t' + ipsw + '---' + str (type (ex).__name__) + '---' + str (ex.args),end='')


if __name__ == '__main__':
    ip_hcl_cisco = []
    ip_hcl_HP=[]
    ip_hcl_HPE=[]
    pid_sw = []
    ip_dict={}
    code = []
    _sw = []
    ou =[]
    amdb = get_inv_solarwinds_live ( )

    for ipam in amdb : #Getting Entire IPAM Code database here and appending in list
        code.append(ipam['IPAM_Code'])

    while True:
        ipam_code = input( f'Please enter IPAM code of site: ' ).upper( ) #Asking user Input for IPAM Code
        if ipam_code in code :
            print( 'IPAM code found, running the program further' )
            user_input(ipam_code)
            # pprint(pid_sw)
            break
        else :
            print ( 'Ipam code not found, please recheck the IPAM code' )


    # wb = xlrd.open_workbook( 'patch_1.xlsx' )
    # sh = wb.sheet_by_index( 0 )
    # a = sh.nrows
    # b = sh.ncols
    #
    # patch_dict = { }
    # patch_list = [ ]
    #
    # for i in range( 1 , a ) :
    #     q = sh.cell_value( i , 0 )
    #     z = sh.cell_value( i , 3 )
    #
    #     patch_dict = { 'model_number' : q , 'newer_version' : z }
    #     patch_list.append( patch_dict )
    #
    ab=openpyxl.Workbook()
    sheet1=ab.active
    #pprint (ou)
    item1 = ou[0]

    x=1
    for k in item1.keys():
        sheet1.cell(row = 1,column = x).value=k
        x=x + 1
    y=2
    for items in ou:
        z=1
        for g in items.values():
            # if "AIR" in g:
            # if type(g) is list:
            #     t='//'.join(g)
            #     sheet1.cell(row = y,column = z).value=t
            # else:
            sheet1.cell(row = y,column = z).value=g

            z=z + 1
        y=y + 1
    total_WAP_Count =  sheet1.max_row - 1
    print (total_WAP_Count)
    # sheet1.cell(row = 1, column = 5).value = 'Version_comparison'
    filename=f'{ipam_code}.xlsx'
    ab.save(filename)
    ab.close()
    # wb1=openpyxl.load_workbook(filename)
    # sheet1=wb1.active

    # o = sheet1.max_row + 1
    #
    # for ipss in range (2, o):
    #
    #
    #     value1 = sheet1.cell(row = ipss, column = 2).value
    #     value_version = sheet1.cell(row = ipss, column = 3).value
    #     # print (value1)
    #     for item in range (len(patch_list)):
    #
    #         if  patch_list[item]['model_number'] in value1:
    #
    #             sheet1.cell(row = ipss, column = 4).value = patch_list[item]['newer_version']
    #
    #             if    patch_list[item]['newer_version'] == "Replace":
    #                 sheet1.cell(row = ipss,column = 5).value = "Replace"
    #
    #             elif    value_version == patch_list[item]['newer_version']:
    #                 sheet1.cell(row = ipss,column = 5).value = "Match"
    #
    #             else:
    #                 sheet1.cell(row = ipss,column = 5).value = "Upgrade Required"

    #
    # wb1.save(filename)
    # wb1.close()



