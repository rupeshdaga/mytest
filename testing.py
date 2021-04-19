import re
import sys
import openpyxl
from pprint import pprint


filename = 'patchmgmtdata.xlsx'
patchmgmtdict = {}
try:
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    sercol = 2
    firmcol = 3
    filecol = 4
    sizecol = 5
    md5col = 6
    for curcol in range(1, ws.max_column + 1):
        if 'Switch Series' in ws.cell(1, curcol).value:
            sercol = curcol
        if 'Firmware Version' in ws.cell(1, curcol).value:
            firmcol = curcol
        if 'File Name' in ws.cell(1, curcol).value:
            filecol = curcol
        if 'File Size' in ws.cell(1, curcol).value:
            sizecol = curcol
        if 'File MD5' in ws.cell(1, curcol).value:
            md5col = curcol

    for currow in range(2, ws.max_row + 1):
        patchmgmtdict[ws.cell(currow, 1).value] = dict()
        patchmgmtdict[ws.cell(currow, 1).value]['SERIES'] = ws.cell(currow, sercol).value
        patchmgmtdict[ws.cell(currow, 1).value]['STD_VER'] = ws.cell(currow, firmcol).value
        patchmgmtdict[ws.cell(currow, 1).value]['STD_FILE_NAME'] = ws.cell(currow, filecol).value
        patchmgmtdict[ws.cell(currow, 1).value]['STD_FILE_SIZE'] = ws.cell(currow, sizecol).value
        patchmgmtdict[ws.cell(currow, 1).value]['STD_FILE_MD5'] = ws.cell(currow, md5col).value
        pprint (patchmgmtdict)
except:
    print('Error opening Patch MGMT data sheet')
    sys.exit()




