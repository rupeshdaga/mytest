import requests
import os
import openpyxl
import login


def get_inv_solarwinds_live(uname=login.username, pasw=login.password, solarwinds_url='https://solarwinds.mdlz.com:17778/SolarWinds/InformationService/v3/Json/Query?query=SELECT%20OrionNodes.NodeId,%20OrionNodes.CustomProperties.Region,%20OrionNodes.CustomProperties.Sub_Region,%20OrionNodes.CustomProperties.Country,%20OrionNodes.CustomProperties.Site,%20OrionNodes.CustomProperties.IPAM_Code,%20OrionNodes.Caption,%20OrionNodes.IP_Address,%20OrionNodes.Status,%20OrionNodes.CustomProperties.Network_Function,%20OrionNodes.CustomProperties.Site_Type,%20OrionNodes.CustomProperties.Site_Operating_Hours,%20OrionNodes.CustomProperties.Site_Metal_Rating,%20OrionNodes.CustomProperties.ManagedBy,%20OrionNodes.SysName,%20OrionNodes.Location,%20OrionNodes.Contact,%20OrionNodes.MachineType,%20OrionNodes.Vendor,%20OrionNodes.LastBoot,%20OrionNodes.IOSVersion%20From%20Orion.Nodes%20AS%20OrionNodes'):
    print('Attempting to download data from SolarWinds...')
    r = requests.get(solarwinds_url, auth=(uname, pasw), verify=False)
    if r.status_code == 200:
        print('Got data from SolarWinds!')
        jas = r.json()
        r.close()
        output = jas.get('results', [])

        wb = openpyxl.load_workbook('example5.xlsx')
        sheet1 = wb.active

        item1 = output[0]
        x = 1
        for k in item1.keys():
            sheet1.cell(row=1, column=x).value = k
            x = x + 1
        y = 2
        for items in output:
            z = 1
            for b in items.values():
                print (b)
                if type(b) is list:
                    q = '//'.join(b)
                    sheet1.cell(row=y, column=z).value = q
                else:
                    sheet1.cell(row=y, column=z).value = b
                z = z + 1
            y = y + 1

        wb.save('example6.xlsx')
    else:
        print(f"Something went wrong... Couldn't load data from SolarWinds. Status Code: {r.status_code}")
    return []

get_inv_solarwinds_live()




