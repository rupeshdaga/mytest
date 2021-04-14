from pprint import pprint
import login
from netmiko import ConnectHandler
import concurrent.futures
import queue
import requests
import openpyxl
import xlrd
import os
from pathlib import Path

class Switch:
    def show_version(self, ip, version, flasha, icode, inv, pipe):
        global model_sup
        global file_name
        abc = 2

        for retry in range (6):
            try:
                my_device={'host': ip , 'username': login.username , 'password': login.password ,
                           'secret': login.password , 'device_type': 'cisco_ios' , 'global_delay_factor': abc}
                rupesh=ConnectHandler ( **my_device )
                prompt = rupesh.find_prompt()
                output = rupesh.send_command(version , expect_string = prompt, use_textfsm = True, delay_factor = 5)
                boot_dir = rupesh.send_command (flasha , expect_string = prompt, use_textfsm = True, delay_factor = 5)

                my_inv = rupesh.send_command (inv, expect_string = prompt, use_textfsm = True, delay_factor = 5)

                for i in range ( len ( my_inv ) ) :
                    if my_inv[i]['pid']=='VS-SUP2T-10G' or my_inv[i]['pid']=='WS-X45-SUP6L-E' or my_inv[i][
                        'pid']=='WS-X45-SUP8L-E' or my_inv[i]['pid']=='WS-X4013+' or my_inv[i]['pid']=='WS-X45-SUP6-E' or\
                            my_inv[i]['pid']=='WS-X45-SUP7-E' or my_inv[i]['pid']=='WS-X45-SUP7L-E':
                        model_sup = output [ 0 ] [ 'hardware' ][0] +  '(' +  my_inv [ i] [ 'pid' ] + ')'

                        break

                    elif 'SUP' not in my_inv[i]['pid']:
                        model_sup = output[0]['hardware'][0]
                for i in range (len(boot_dir)):
                    if boot_dir[i]['name'] in new_ios_file:
                        file_name = 'Present'
                        break
                    else:
                        file_name = "Not Present"

                sw_dict = { 'IPAM_Code': icode , 'IP_Address': ip , "Model": model_sup ,
                            "Hostname": output[0]['hostname'] , "IOS_version": output[0]['version'] ,"Running_Image" : output[0]['running_image'].strip('/'),
                            "Stack_Count": (len ( output[0]['hardware'] )) ,
                            "Boot_Dir": boot_dir[0]['file_system'].strip ( '/' ) , "New_File_In_Flash" : file_name,
                            "Free_Memory": boot_dir[0]['total_free']
                            }

                pipe.put ( sw_dict)
                rupesh.disconnect()
                return

            except Exception as ex:
                print ( f'\n\t ! Login into {ip} failed.' , end = '' )
                print ( '\n\t' + ip + '---' + str ( type ( ex ).__name__ ) + '---' + str ( ex.args ) , end = '' )
                abc= abc * 4



def get_inv_solarwinds_live(uname=login.username , pasw=login.password ,
                                solarwinds_url='https://solarwinds.mdlz.com:17778/SolarWinds/InformationService/v3/Json/Query?query=SELECT%20OrionNodes.NodeId,%20OrionNodes.CustomProperties.Region,%20OrionNodes.CustomProperties.Sub_Region,%20OrionNodes.CustomProperties.Country,%20OrionNodes.CustomProperties.Site,%20OrionNodes.CustomProperties.IPAM_Code,%20OrionNodes.Caption,%20OrionNodes.IP_Address,%20OrionNodes.Status,%20OrionNodes.CustomProperties.Network_Function,%20OrionNodes.CustomProperties.Site_Type,%20OrionNodes.CustomProperties.Site_Operating_Hours,%20OrionNodes.CustomProperties.Site_Metal_Rating,%20OrionNodes.CustomProperties.ManagedBy,%20OrionNodes.SysName,%20OrionNodes.Location,%20OrionNodes.Contact,%20OrionNodes.MachineType,%20OrionNodes.Vendor,%20OrionNodes.LastBoot,%20OrionNodes.IOSVersion%20From%20Orion.Nodes%20AS%20OrionNodes'):
    print ( 'Attempting to download data from SolarWinds...' )
    r=requests.get ( solarwinds_url , auth = (uname , pasw) , verify = False )
    if r.status_code == 200:
        print ( 'Got data from SolarWinds!' )
        jas=r.json ( )
        r.close ( )
        return jas.get ( 'results' , [] )
    else:
        print ( f"Something went wrong... Couldn't load data from SolarWinds. Status Code: {r.status_code}" )
    return []
if __name__ == '__main__':

    amdb=get_inv_solarwinds_live ( )
    code = []
    ip_hcl_cisco = []
    ip_hcl_HP=[]
    ip_hcl_HPE=[]
    sw_dict={}
    qdata=[]
    md5data = []
    new_ios_file = []


    wb = openpyxl.load_workbook ( 'patch_md5.xlsx' ) # Opening the Main Patch Database file
    sh = wb.active
    a = sh.max_row + 1
    b = sh.max_column + 1

    for ios in range (2, a):
        new_ios_file.append(sh.cell(row = ios, column=4).value)

    # pprint (new_ios_file)
    for ipam in amdb:  # Getting Entire IPAM Code database here and appending in list
        code.append ( ipam ['IPAM_Code'] )

    while True:
        ipam_code=input ( f'Please enter IPAM code of site: ' ).upper ( )  # Asking user Input for IPAM Code
        if ipam_code in code:
            print ( 'IPAM code found, running the program further' )
            for item in amdb:
                if item ['IPAM_Code'] == ipam_code and 'Switch' in item ['Network_Function'] and item ['Vendor'] == 'H3C':
                    ip_hcl_HP.append ( item ['IP_Address'] )

            for item in amdb:
                if item ['IPAM_Code'] == ipam_code and "Switch" in item ['Network_Function'] and item ['Vendor'] == "Cisco":
                    ip_hcl_cisco.append ( item ['IP_Address'] )

            for item in amdb:
                if item ['IPAM_Code'] == ipam_code and "Switch" in item ['Network_Function'] and item ['Vendor'] == "HPE":
                    ip_hcl_HPE.append ( item ['IP_Address'] )
            break
        else:
            print ( 'IPAM code not found, please recheck the IPAM code')

    ra=Switch ( )
    xyz = queue.Queue()
    with concurrent.futures.ThreadPoolExecutor ( max_workers = 25 ) as mainexecutor:
        for myip in ip_hcl_cisco:
           mainexecutor.submit (ra.show_version, myip, 'show version', 'dir', ipam_code, 'show inventory', xyz)

    while not xyz.empty ( ):
        qdata.append (xyz.get())

    ab = openpyxl.Workbook ( )
    sheet1 = ab.active
    item1 = qdata[0]
    x = 1
    for k in item1.keys ( ) :
        sheet1.cell (row = 1, column = x ).value = k
        x = x + 1
    y = 2
    for items in qdata :
        z = 1
        for g in items.values ( ) :
            sheet1.cell (row = y , column = z).value = g
            z = z + 1
        y = y + 1



    colx = sheet1.max_column
    rowx = sheet1.max_row + 1


    colx += 1
    sheet1.cell ( row = 1 , column =  colx).value = "New_IOS_Version"
    colx += 1

    sheet1.cell ( row = 1 , column = colx ).value = 'New_IOS_Name'
    colx += 1
    sheet1.cell ( row = 1 , column = colx ).value = "Version_comparison"
    colx += 1
    sheet1.cell ( row = 1 , column = colx ).value = "MD5_Value"
    colx+= 1
    sheet1.cell ( row = 1 , column = colx ).value = "Memory_Status"
    colx += 1
    sheet1.cell (row = 1,column = colx).value = "Command "
    patch_dict = { }
    patch_list = []

    for i in range ( 2 , a ):
        model_c = sh.cell( i , 1 ).value
        # print (model_c)
        New_IOS = sh.cell( i , 3 ).value
        MD5_Value = sh.cell(i, 6).value
        New_IOS_Name = sh.cell(i, 4).value
        File_Size = sh.cell(i, 5).value
        patch_dict = { 'model_number': model_c , 'newer_version': New_IOS , 'New_IOS' : New_IOS_Name, 'MD5_value' : MD5_Value, 'File_Size' : File_Size}
        patch_list.append ( patch_dict )



    for ipss in range (2, rowx ):

        value1_model = sheet1.cell(row = ipss, column = 3).value
        value_version = sheet1.cell(row = ipss, column = 5).value
        free_space = sheet1.cell ( row = ipss , column = 10 ).value   #

        for item in range (len(patch_list)):

            try:
                if  patch_list[item]['model_number'] in value1_model:
                    sheet1.cell(row = ipss, column = 11).value = patch_list[item]['newer_version']
                    sheet1.cell(row = ipss, column = 12).value = patch_list[item]['New_IOS']
                    sheet1.cell ( row = ipss , column = 14 ).value = patch_list[item]['MD5_value']

                    if  patch_list[item]['newer_version'] == "Replace":
                        sheet1.cell(row = ipss,column = 13).value = "Replace"
                        sheet1.cell ( row = ipss , column = 15 ).value = "NA"


                    if  value_version == patch_list[item]['newer_version']:
                        sheet1.cell(row = ipss,column = 13).value = "Match"
                        sheet1.cell (row = ipss,column = 15).value = "NA"


                    if  value_version !=  patch_list[item]['newer_version']:
                        sheet1.cell(row = ipss,column = 13).value = "Upgrade Required"

                        if int(free_space)  > int(patch_list[item]['File_Size']):
                            sheet1.cell ( row = ipss , column = 15 ).value = "Memory Available"

                        if int(free_space) < int (patch_list[item]['File_Size']):
                            sheet1.cell ( row = ipss , column = 15 ).value = "Cleanup Required"
            except:
                continue



    rowx = sheet1.max_row + 1

    for abc in range(2,rowx):
        if sheet1.cell (row = abc,column = 5).value == '03.06.07E':
            sheet1.cell (row = abc,column = 16).value = 'yes'

            # command = f'software install file {sheet1.cell (row = abc,column = 8).value}' \
            #     f'{sheet1.cell (row = abc,column = 12).value} new force'
            # sheet1.cell (row = abc,column = 16).value = command
                # print (f" Switch with Ip {sheet1.cell(row=abc,column=2).value} requires command {command}")

    myfile = f'{ipam_code}.xlsx'
    ab.save (myfile)

        # elif 'C3650' or 'C3850' or 'C9300' not in sheet1.cell(row=abc,column=3).value and sheet1.cell(row=abc,column=14).value\
        #         =='Memory Available' and sheet1.cell(row=abc,column=12).value=='Upgrade Required':


