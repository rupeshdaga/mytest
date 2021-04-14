import re
import sys
import openpyxl


def patchmgmtdata(filename='patchmgmtdata.xlsx'):
    """
    Opens MDLZ standard patchmanagement sheet and returns data in dictionary form
    @param filename: file name of excel sheet
    @return: dictionary of MDLZ patch management
    """
    patchmgmtdict = dict()
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        sercol = 2
        firmcol = 3
        filecol = 4
        sizecol = 5
        md5col = 6
        for curcol in range(1, ws.max_column + 1):
            if 'Switch Series' in ws.cell(1, curcol).value:
                sercol = curcol
            if 'Firmware Version' in ws.cell(1, curcol).value:
                firmcol = curcol
            if 'File Name' in ws.cell(1, curcol).value:
                filecol = curcol
            if 'File Size' in ws.cell(1, curcol).value:
                sizecol = curcol
            if 'File MD5' in ws.cell(1, curcol).value:
                md5col = curcol

        for currow in range(2, ws.max_row + 1):
            patchmgmtdict[ws.cell(currow, 1).value] = dict()
            patchmgmtdict[ws.cell(currow, 1).value]['SERIES'] = ws.cell(currow, sercol).value
            patchmgmtdict[ws.cell(currow, 1).value]['STD_VER'] = ws.cell(currow, firmcol).value
            patchmgmtdict[ws.cell(currow, 1).value]['STD_FILE_NAME'] = ws.cell(currow, filecol).value
            patchmgmtdict[ws.cell(currow, 1).value]['STD_FILE_SIZE'] = ws.cell(currow, sizecol).value
            patchmgmtdict[ws.cell(currow, 1).value]['STD_FILE_MD5'] = ws.cell(currow, md5col).value
        return patchmgmtdict
    except:
        print('Error opening Patch MGMT data sheet')
        sys.exit()


def check_patch_mgmt_for_version(sh_ver, sh_inv, patchmgmtdict, chassis_model_list=('6807', '4503', '4506', '4507', '4509', '6509')):
    """
    Accepts output from switch and returns patch management status
    @param sh_ver: Show version output parsed by TextfSM
    @param sh_inv: Show Inventory output parsed by TextFSM
    @param patchmgmtdict: patchmanagement dictionary generated by patchmgmtdata function
    @param chassis_model_list: list of chassis
    @return: dictionary with various important details
    """
    localpmdata = {'PATCH_MGMT_STATUS': '',
                   'NOT_ON_PM_FLAG': False,
                   'PATCH_MGMT_VER': '',
                   'PATCH_MGMT_FILE': '',
                   'PATCH_MGMT_SIZE': '',
                   'PATCH_MGMT_MD5': '',
                   }

    swver = ''
    swmod = ''

    for ver in sh_ver:
        swver = ver.get('version', '')
        for hw in ver.get('hardware', []):
            swmod = hw
            break
        break

    re_checklic = re.compile(r'[A-Za-z]-[ELSels]$')
    re_removelic = re.compile(r'-[ELSels]$')
    re_fixver = re.compile(r'\.E$')

    chassis_flag = False
    for chassis_model in chassis_model_list:
        if chassis_model in swmod:
            chassis_flag = True
            break

    # if '6807' in swmod or '4503' in swmod or '4506' in swmod or '4507' in swmod or '4509' in swmod or '6509' in swmod:
    if chassis_flag:
        for i in sh_inv:
            if 'supervisor' in i['name'].lower() or 'supervisor' in i['descr'].lower():
                swmodsp = swmod + '(' + i['pid'] + ')'
                break
        else:
            print('SUP engine information not found!')
            swmodsp = swmod
    else:
        if re_checklic.search(swmod) != None:
            swmodsp = re_removelic.sub('', swmod)
        else:
            swmodsp = swmod

    if patchmgmtdict.get(swmodsp) is None:
        localpmdata['PATCH_MGMT_STATUS'] = f'Please update Patch MGMT with model "{swmodsp}"'
        return localpmdata
    else:
        if patchmgmtdict[swmodsp]['STD_VER'] == 'Replace':
            localpmdata['PATCH_MGMT_STATUS'] = 'EOL/EOS - Replace Immediately'
            return localpmdata
        elif re_fixver.sub('E', swver) == patchmgmtdict[swmodsp]['STD_VER']:
            localpmdata['PATCH_MGMT_STATUS'] = 'On MDLZ Standard'
            return localpmdata
        else:
            localpmdata['PATCH_MGMT_STATUS'] = 'NOT on MDLZ approved IOS'
            localpmdata['PATCH_MGMT_VER'] = patchmgmtdict[swmodsp]['STD_VER']
            localpmdata['PATCH_MGMT_FILE'] = patchmgmtdict[swmodsp]['STD_FILE_NAME']
            localpmdata['PATCH_MGMT_SIZE'] = patchmgmtdict[swmodsp]['STD_FILE_SIZE']
            localpmdata['PATCH_MGMT_MD5'] = patchmgmtdict[swmodsp]['STD_FILE_MD5']
            localpmdata['NOT_ON_PM_FLAG'] = True
            return localpmdata


def check_file_md5(fname):
    """
    This function accepts file and returns MD5 checksum
    @param fname: file location
    @return: MD5 string
    """
    import hashlib
    hash_md5 = hashlib.md5()
    with open(fname, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.digest()


if __name__ == '__main__':
    input('Please execute auto_patch.py')