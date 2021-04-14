import openpyxl
import pprint

import datetime

from datetime import datetime

cosmos = "*************************Cosmos Tarders and Stationers************************"

print (cosmos)

wb = openpyxl.Workbook()
sheet1 = wb.active
sheet1 = wb['Sheet']

Amount = list()
amt_gst = list()
Amount_GST = 0
cosmos_dict={}
cosmos_list = []

def billing():

    abc = int(input("Serial N0: "))
    Goods = input("Description of Goods: ")
    HSN_SAC = input("HSN_SAC value: ")
    Quantity = int(input("Quantity : "))
    Rate_1 = float(input("Rate per Quantity: "))
    per = input('per: ')
    sgst = float(input("SGST_% : "))
    cgst = float (input("CGST_%: "))
    Amount_GST = float((Quantity * Rate_1 * sgst) / 100)

    Total_Amount = float(Quantity*Rate_1) + (Amount_GST + Amount_GST)
    Amount.append(Total_Amount)
    amt_gst.append(Amount_GST)

    cosmos_dict = {'Serial Number': abc ,'Goods_1': Goods, "HSN/SAC": HSN_SAC, "QTY": Quantity, "rate": Rate_1,'Per': per, "sgst_%": sgst, "Amount_SGST": Amount_GST,"cgst_%": cgst, "Amount_CGST": Amount_GST, "Total": Total_Amount}
    cosmos_list.append(cosmos_dict)

    Answer = input ("Do you wish to continue (yes or no): ")
    if Answer == 'yes':
        billing()
    else:
        xyz = len(Amount) + 3 + 3
        item1 = cosmos_list[0]
        x = 1
        for k in item1.keys():
            sheet1.cell(row=3, column=x).value = k
            x = x + 1
        sheet1.cell(row = 1, column = 1).value = "Date of BIll:"
        sheet1.cell(row = 2, column = 1).value = 'Customer Name'
        sheet1.cell(row = 1, column = 3).value = 'GSTIN_Number'
        sheet1.cell(row = 2, column = 2).value  = input("Name of Customer : ")
        sheet1.cell(row = 1, column = 2).value = input("Date of Bill: ")
        y = 4
        for items in cosmos_list:
            z = 1
            for b in items.values():
                sheet1.cell(row=y, column=z).value = b
                z = z + 1
            y = y + 1
            d = 0
            gst = 0
            for i in range (len(Amount)):
                d += Amount[i]
            for k in range(len(amt_gst)):
                gst += amt_gst[k]
            sheet1.cell(row = xyz, column = 7).value = "Total SGST"
            sheet1.cell(row = xyz, column = 8).value = gst
            sheet1.cell(row = xyz, column = 9).value = "Total CGST"
            sheet1.cell(row = xyz - 1, column = 11).value = "Total Amount"
            sheet1.cell(row = xyz, column = 10).value = gst
            sheet1.cell(row= xyz, column = 11).value = d
    wb.save('cosmos_bill.xlsx')
    wb.save('cosmos2.xlsx')

billing()
