import openpyxl
wb1=openpyxl.load_workbook('Link_down.xlsx')
sheet1=wb1.active
wb = openpyxl.load_workbook('IPAM.xlsx')
sh = wb.active
aa = sh.max_row + 1
ipam_list = []
for rr in range (2, aa):
    ipam_list.append(sh.cell(row = rr, column = 1).value)
a = sheet1.max_row + 1
for i in range(2, a) :
    for xx in range(len(ipam_list)) :
        if ipam_list [xx] in sheet1.cell(row = i, column = 1).value :
            sheet1.cell(row = i, column = 2).value = ipam_list [xx]
wb1.save('Link_down.xlsx')
wb1.close()

wb.save('IPAM.xlsx')
wb.close()

