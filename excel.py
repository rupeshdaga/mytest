import openpyxl

work_book = openpyxl.load_workbook('rupesh_abcd.xlsx')

#print(work_book.sheetnames) # Prints out all the sheet available in openpyxl

sheet = work_book['name']  # Assigning first sheet of rupesh_abcd.xlsx to object named sheet

#print(sheet.title) # This prints out the name of sheet, output is name


# print(sheet ['A2'].value) # This will dospaly the content of call A2 in our xlsx file (rupesh_abcd), we can also call this in variable as below

# cell = sheet ['A2'].value # Same as above but called in variable cell

#print (type(cell)) # Return object is string



#print (type(cell)) # Without using value key word, output is "<class 'openpyxl.cell.cell.Cell'>"

#print(dir(cell))


#print(sheet ['A2'].value + ' of age: ' + str(sheet['B2'].value) + ' knows coding: ' + sheet ['C2'].value + ', but he is ' + sheet ['D2'].value + ' today') # Concaneting all cells

#print (sheet.max_row) # Will indicate we are in which particular cell

#print(sheet.cell(row=3, column= 2).value) # Returns a value of specific cell

#print(sheet.max_row) # Specifies maximum number for rows (In our case it is 4 rows)

#print(sheet.max_column) # Specifies maximum number of columns (In our case it is 5 columns)





