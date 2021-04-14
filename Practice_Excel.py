import openpyxl
from pprint import pprint
from netmiko import ConnectHandler
import login
import concurrent.futures


def File_Delete (myip, myver, ruver, myios,bootd):

    my_device = {
        'host' : myip, 'username' : login.username, 'password' : login.password, 'secret' : login.password,
        'device_type' : 'cisco_ios', 'global_delay_factor' : 15
    }
    rupesh = ConnectHandler( **my_device )
    prompt = rupesh.find_prompt( )
    output = rupesh.send_command( 'dir', expect_string=prompt, use_textfsm=True, delay_factor= 10)
    for i in range (len(output)):
        if 'cat3k_caa' in output[i]['name']:
            x.append (output [ i ] [ 'name' ])
    # pprint (x)

    if  myver in x:
        print (f"New Image Already present for switch {myip}")
    else:
        print (f"New Image not found for switch {myip}")

    # elif ruver > 16:
    #     command = f'request platform software package clean switch  all'
    #     print (f" Command for Switch {myip} is {command}")
    # elif ruver < 16:
    #     command = f'software clean'
    #     print (f" Command for Switch {myip} is {command}")



    rupesh.disconnect()



if __name__ == '__main__':

    ab = openpyxl.load_workbook('INTAE.xlsx')

    sh = ab.active
    a = sh.max_row+ 1

    x = []

    with concurrent.futures.ThreadPoolExecutor (max_workers=50) as mainexecutor:
        for v in range ( 2 , a ) :
            if  sh.cell (row=v , column=14).value=='Cleanup Required' :
                ip=sh.cell (row=v , column=2).value
                # print (ip)
                nver=sh.cell (row=v , column=11).value
                # print (f"New version of image is {nver}")
                boot_path = sh.cell(row = v, column=8).value
                runn_ver =  sh.cell(row = v, column=5).value
                IOS_value = sh.cell(row = v, column=10).value
                # print (f"Running_IOS_Version is {runn_ver}")

                mainexecutor.submit (File_Delete, ip , nver, runn_ver, IOS_value, boot_path)
